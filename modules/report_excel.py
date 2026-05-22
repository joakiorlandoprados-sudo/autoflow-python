"""Excel report export utilities."""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import config


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)


def export_excel(df: pd.DataFrame, stats: dict[str, Any], output_path: str) -> None:
    """Export cleaned data and summary statistics to a styled Excel workbook."""

    destination = Path(output_path).expanduser()
    if not destination.is_absolute():
        destination = (Path.cwd() / destination).resolve()
    destination.parent.mkdir(parents=True, exist_ok=True)

    try:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

        data_sheet = workbook.create_sheet("Cleaned Data")
        stats_sheet = workbook.create_sheet("Statistics")
        chart_sheet = workbook.create_sheet("Chart")

        _write_data_sheet(data_sheet, df)
        _write_stats_sheet(stats_sheet, stats)
        _write_chart_sheet(chart_sheet, df)

        workbook.save(destination)
        config.safe_print(f"✅ Excel report saved to {destination}")
    except Exception as error:
        raise RuntimeError(f"Failed to export Excel report: {error}") from error


def _write_data_sheet(worksheet: Any, df: pd.DataFrame) -> None:
    """Populate the cleaned data sheet and apply formatting."""

    worksheet.append([str(column) for column in df.columns])
    _style_header_row(worksheet, 1)

    export_df = df.copy()
    for column in export_df.select_dtypes(include=["datetime64[ns]", "datetimetz"]).columns:
        export_df[column] = export_df[column].dt.strftime("%Y-%m-%d")

    for row in export_df.itertuples(index=False, name=None):
        worksheet.append(list(row))

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")

    _autofit_columns(worksheet)


def _write_stats_sheet(worksheet: Any, stats: dict[str, Any]) -> None:
    """Populate the statistics summary sheet."""

    row_cursor = 1
    worksheet.cell(row=row_cursor, column=1, value="Dataset Overview")
    _style_section_title(worksheet, row_cursor)
    row_cursor += 1

    overview_items = [
        ("Rows", stats["dataset"]["row_count"]),
        ("Columns", stats["dataset"]["column_count"]),
        ("Generated At", stats["dataset"]["generated_at"]),
        ("Numeric Columns", ", ".join(stats["dataset"]["numeric_columns"]) or "None"),
        (
            "Categorical Columns",
            ", ".join(stats["dataset"]["categorical_columns"]) or "None",
        ),
        ("Date Columns", ", ".join(stats["dataset"]["date_columns"]) or "None"),
    ]
    for label, value in overview_items:
        worksheet.cell(row=row_cursor, column=1, value=label)
        worksheet.cell(row=row_cursor, column=2, value=value)
        row_cursor += 1

    row_cursor += 1
    worksheet.cell(row=row_cursor, column=1, value="Numeric Summary")
    _style_section_title(worksheet, row_cursor)
    row_cursor += 1

    numeric_header = ["Column", "Min", "Max", "Mean", "Median", "Std", "Total", "Missing"]
    for index, heading in enumerate(numeric_header, start=1):
        worksheet.cell(row=row_cursor, column=index, value=heading)
    _style_header_row(worksheet, row_cursor)
    row_cursor += 1

    if stats["numeric"]:
        for column, values in stats["numeric"].items():
            worksheet.append(
                [
                    column,
                    values["min"],
                    values["max"],
                    values["mean"],
                    values["median"],
                    values["std"],
                    values["total"],
                    values["missing"],
                ]
            )
    else:
        worksheet.cell(row=row_cursor, column=1, value="No numeric columns available.")
        row_cursor += 1

    row_cursor = worksheet.max_row + 2
    worksheet.cell(row=row_cursor, column=1, value="Categorical Highlights")
    _style_section_title(worksheet, row_cursor)
    row_cursor += 1

    categorical_header = ["Column", "Unique Values", "Top 5 Values"]
    for index, heading in enumerate(categorical_header, start=1):
        worksheet.cell(row=row_cursor, column=index, value=heading)
    _style_header_row(worksheet, row_cursor)
    row_cursor += 1

    if stats["categorical"]:
        for column, values in stats["categorical"].items():
            top_values = ", ".join(
                f"{item['value']} ({item['count']})" for item in values["top_values"]
            )
            worksheet.append([column, values["unique_values"], top_values])
    else:
        worksheet.cell(row=row_cursor, column=1, value="No categorical columns available.")
        row_cursor += 1

    time_trend = stats.get("time_trend")
    if time_trend:
        row_cursor = worksheet.max_row + 2
        worksheet.cell(row=row_cursor, column=1, value="Time Trend")
        _style_section_title(worksheet, row_cursor)
        row_cursor += 1
        trend_items = [
            ("Date Column", time_trend["date_column"]),
            ("Metric Column", time_trend["metric_column"] or "Record Count"),
            ("Granularity", time_trend["granularity"]),
            ("Start Date", time_trend["start_date"]),
            ("End Date", time_trend["end_date"]),
            ("Summary", time_trend["summary"]),
        ]
        for label, value in trend_items:
            worksheet.cell(row=row_cursor, column=1, value=label)
            worksheet.cell(row=row_cursor, column=2, value=value)
            row_cursor += 1

    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _autofit_columns(worksheet)


def _write_chart_sheet(worksheet: Any, df: pd.DataFrame) -> None:
    """Create a simple bar chart based on a relevant numeric column."""

    chart_source = _build_chart_source(df)
    if chart_source.empty:
        worksheet["A1"] = "No suitable numeric and categorical data found for charting."
        return

    worksheet.append(list(chart_source.columns))
    _style_header_row(worksheet, 1)
    for row in chart_source.itertuples(index=False, name=None):
        worksheet.append(list(row))

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = f"{chart_source.columns[1]} by {chart_source.columns[0]}"
    chart.y_axis.title = chart_source.columns[0]
    chart.x_axis.title = chart_source.columns[1]

    data_reference = Reference(
        worksheet, min_col=2, min_row=1, max_row=worksheet.max_row, max_col=2
    )
    category_reference = Reference(worksheet, min_col=1, min_row=2, max_row=worksheet.max_row)
    chart.add_data(data_reference, titles_from_data=True)
    chart.set_categories(category_reference)
    chart.height = 8
    chart.width = 18

    worksheet.add_chart(chart, "D2")
    _autofit_columns(worksheet)


def _build_chart_source(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregate a small chart dataset from the best available columns."""

    numeric_columns = [
        column
        for column in df.select_dtypes(include=["number"]).columns
        if not pd.api.types.is_bool_dtype(df[column])
    ]
    categorical_columns = list(df.select_dtypes(include=["object", "string", "category"]).columns)
    date_columns = list(df.select_dtypes(include=["datetime64[ns]", "datetimetz"]).columns)

    if not numeric_columns:
        return pd.DataFrame()

    metric_column = _pick_chart_metric(numeric_columns)

    if categorical_columns:
        group_column = categorical_columns[0]
        chart_df = (
            df.groupby(group_column, dropna=False)[metric_column]
            .sum()
            .sort_values(ascending=False)
            .head(10)
            .reset_index()
        )
        chart_df[group_column] = chart_df[group_column].fillna("Missing").astype(str)
        return chart_df

    if date_columns:
        date_column = date_columns[0]
        chart_df = (
            df.dropna(subset=[date_column])
            .groupby(pd.Grouper(key=date_column, freq="MS"))[metric_column]
            .sum()
            .reset_index()
        )
        chart_df[date_column] = chart_df[date_column].dt.strftime("%Y-%m")
        return chart_df

    return pd.DataFrame()


def _pick_chart_metric(numeric_columns: list[str]) -> str:
    """Pick a numeric column that is likely to be meaningful in a chart."""

    priorities = ("net_revenue", "total_revenue", "revenue", "sales", "amount", "units_sold")
    for keyword in priorities:
        for column in numeric_columns:
            if keyword in column.lower():
                return column
    return numeric_columns[0]


def _style_header_row(worksheet: Any, row_number: int) -> None:
    """Apply table header styles to a worksheet row."""

    for cell in worksheet[row_number]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER


def _style_section_title(worksheet: Any, row_number: int) -> None:
    """Apply section heading styles to a single worksheet row."""

    title_cell = worksheet.cell(row=row_number, column=1)
    title_cell.font = Font(bold=True, color="1F1F1F")
    title_cell.fill = SECTION_FILL
    title_cell.border = THIN_BORDER
    title_cell.alignment = Alignment(horizontal="left")


def _autofit_columns(worksheet: Any) -> None:
    """Size worksheet columns based on visible content."""

    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 35)
